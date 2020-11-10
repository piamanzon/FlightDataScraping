'''
Created on Oct 5, 2020

@author: pia.manzon
'''
from openpyxl import workbook #pip install openpyxl
from openpyxl import load_workbook

from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from datetime import datetime
from datetime import date


class Excel():
    def __init__(self,excelPath):
        self.filePath = excelPath
        self.workbook = load_workbook(excelPath)
        self.worksheet = self.workbook.active
        self.myThinBorder = Border(top=Side(border_style='thin', color='000000'))
        self.addStyle()
       
    def addStyle(self):
        if(not self.styleExists("RawDataStyle")):
            self.rawdataStyle = NamedStyle(name = "RawDataStyle", font = Font(name = 'Arial', size =10), alignment = Alignment(horizontal = 'left', vertical = 'bottom'))
            self.workbook.add_named_style(self.rawdataStyle)
        if(not self.styleExists("ChartDataStyle")):
            self.chartDataStyle = NamedStyle(name = "ChartDataStyle", font = Font(name = 'Arial', size =10), alignment = Alignment(horizontal = 'center', vertical = 'bottom'))
            self.workbook.add_named_style(self.chartDataStyle)
        if(not self.styleExists("CellStyle")):
            self.cellStyle = NamedStyle(name = "CellStyle", number_format = 'd-mmm-yy', font = Font(name = 'Arial', size =10), alignment = Alignment(horizontal = 'left', vertical = 'bottom'))
            self.workbook.add_named_style(self.cellStyle)
      
    def styleExists(self,styleName):
        result = self.workbook.named_styles
         
        if (styleName in result):
            return True
        return False
        
    def getMaxRow(self,columnName):
        maxRow = max((c.row for c in self.worksheet[columnName] if c.value is not None))
        return maxRow
    
    def setBorder(self,maxRow,columnStart):
        for index in range(columnStart,columnStart+3):
            self.worksheet.cell(row = maxRow, column = index).border = self.myThinBorder 
        
    def inputNewData (self,maxRow,flightDataList,city):       
        newMaxRow = maxRow + 1
        columnStart = 0
        borderRow = newMaxRow

        if (city == "Vancouver"):
            columnStart = 3
                 
        else:
            columnStart = 7
              
#Update Flight Info
#===============================================================================
        for index in range(len(flightDataList)):
            tempCell =  self.worksheet.cell(row=newMaxRow, column = columnStart)
            tempCell.value = datetime.strptime(flightDataList[index]['Date'], '%m/%d/%Y')
            tempCell.style = "CellStyle"
            tempCell =  self.worksheet.cell(row=newMaxRow, column = (columnStart+1))
            tempCell.value = flightDataList[index]['Flight']
            tempCell.style = "RawDataStyle"
            tempCell = self.worksheet.cell(row=newMaxRow, column = (columnStart+2))
            tempCell.value = flightDataList[index]['From']
            tempCell.style = "RawDataStyle"
            newMaxRow+=1   
        self.setBorder(borderRow, columnStart)

#Update Daily Summary
#===============================================================================  
    def updateDailySummary(self,flightVancouverList,flightTorontoList):
        dateToday = datetime.strptime(flightVancouverList[0]['Date'], '%m/%d/%Y')
        columnStart = 12 
        newMaxRow = self.getMaxRow('l') + 1
        tempCell = self.worksheet.cell(row=newMaxRow, column = columnStart)
        tempCell.value = dateToday
        tempCell.style = "CellStyle"
        rowC = 'C'+str(self.getMaxRow('C'))
        print(rowC)
       
        rowL = 'L' + str(newMaxRow)
        print (rowL)   
        vanCell = 'M' + str(newMaxRow)
        torCell = 'N' + str(newMaxRow)
        self.worksheet[vanCell] = "=COUNTIF(C13:" + rowC + "," + rowL + ")"
        rowC = 'G'+str(self.getMaxRow('G'))
        self.worksheet[torCell] = "=COUNTIF(G13:" + rowC + "," + rowL + ")"
        #=======================================================================
        # tempCell = self.worksheet.cell(row=newMaxRow, column = (columnStart+1))
        # tempCell.value = len(flightVancouverList)
        # tempCell.style = "ChartDataStyle"
        # tempCell =   self.worksheet.cell(row=newMaxRow, column = (columnStart+2)) 
        # tempCell.value = len(flightTorontoList)
        # tempCell.style = "ChartDataStyle"
        #=======================================================================
        
        
   
    def updateWeeklySummary(self):
        x = 133
        while x <=(137):
            dateMaxRow = self.getMaxRow('L') #Max weekly date row
            dateMaxRow = x
            #print(dateMaxRow)
            #dateToday = date.today()
            dateToday = self.worksheet.cell(row=dateMaxRow, column = 12).value
            dateToday = dateToday[:7] + "20" + dateToday[7:]
                    # dateLastWeek = datetime.strptime(formattedDate, "%d-%b-%Y")
            dateToday= datetime.strptime(dateToday, "%d-%b-%Y")
            print("Date today is ", dateToday)
            currentWeek = dateToday.isocalendar()[1]
            print("Current week is ", currentWeek)
            #Weekly Date
           # weekNumToday = datetime.date(dateToday).strftime("%V")
            #
            #dateMaxRow = 16
            tempCellVal = self.worksheet.cell(row=dateMaxRow-1, column = 12).value
            print(tempCellVal)
            formattedDate = tempCellVal[:7] + "20" + tempCellVal[7:]
            dateLastWeek = datetime.strptime(formattedDate, "%d-%b-%Y")
            print("Date last week is ", tempCellVal)
            #print(formattedDate)
            weekBefore = dateLastWeek.isocalendar()[1]
            print("Week before is ", weekBefore)
            
            flightMaxRow = self.getMaxRow('M') 
            flightMaxRow = x
            tempCell = self.worksheet.cell(row=flightMaxRow, column = 13)
            tempCellYYZ = self.worksheet.cell(row=flightMaxRow, column = 14)
            dataList = []
            dataListYYZ = []
            average = 0
            averageYYZ = 0
            index = 0
            if(currentWeek != weekBefore):
                average = tempCell.value
                averageYYZ = tempCellYYZ.value
                newMaxRow = self.getMaxRow('O')+1
                tempCell = self.worksheet.cell(row=newMaxRow, column = 15)
                #tempCell.value= self.worksheet.cell(row=x, column = 12).value.strftime('%d-%b-%y').lstrip("0").replace(" 0", " ")
                tempCell.value= self.worksheet.cell(row=x, column = 12).value
                tempCell.style = "RawDataStyle"
              #  tempCell.style = "RawDataStyle"
                tempCell = self.worksheet.cell(row=newMaxRow, column = 16)
                tempCellYYZ = self.worksheet.cell(row=newMaxRow, column = 17)
                tempCell.value = average
                tempCellYYZ.value = averageYYZ
                tempCell.style = "ChartDataStyle"
                tempCellYYZ.style = "ChartDataStyle"
    #print(average)
            else:
                index = flightMaxRow
                while (currentWeek == weekBefore):
                    tempCell = self.worksheet.cell(row=index, column = 13)
                    tempCellYYZ = self.worksheet.cell(row=index, column = 14)
                    print("Current date flight is ", tempCell.value)
                    dataList.append(tempCell.value)
                    dataListYYZ.append(tempCellYYZ.value)
                    index= index -1
                    currentWeek = weekBefore
                    print("Current week is ", currentWeek)
                    tempCellVal = self.worksheet.cell(row=index-1, column = 12).value
                    dateLastWeek = tempCellVal
                    #===========================================================
                    # formattedDate = tempCellVal[:7] + "20" + tempCellVal[7:]
                    # dateLastWeek = datetime.strptime(formattedDate, "%d-%b-%Y")
                    #===========================================================
                    print("Date last week is ", dateLastWeek)
                    weekBefore = dateLastWeek.isocalendar()[1] 
                    print("Week before is ", weekBefore)
                tempCell = self.worksheet.cell(row=index, column = 13)
                print("Current date flight is ", tempCell.value)
                dataList.append(tempCell.value)
                average = round((sum(dataList)/len(dataList)),0) 
                dataListYYZ.append(tempCellYYZ.value)
                averageYYZ = round((sum(dataListYYZ)/len(dataListYYZ)),0)
                print ("Average is ", average)
                newMaxRow = self.getMaxRow('O')
                tempCell = self.worksheet.cell(row=newMaxRow, column = 15)
                #tempCell.value= self.worksheet.cell(row=index, column = 12).value.strftime('%d-%b-%y').lstrip("0").replace(" 0", " ")
                tempCell.value= self.worksheet.cell(row=index, column = 12).value
              #  tempCell.style = "RawDataStyle"
                tempCell = self.worksheet.cell(row=newMaxRow, column = 16)
                tempCell.value = average
                tempCell.style = "ChartDataStyle"
                tempCellYYZ = self.worksheet.cell(row=newMaxRow, column = 17)
                tempCellYYZ.value = averageYYZ
                tempCellYYZ.style = "ChartDataStyle"
            x = x + 1
       # tempCell.style = "ChartDataStyle"
        
        
        #=======================================================================
        # newMaxRow = self.getMaxRow('O') + 1 #Max weekly date row
        # tempCell = self.worksheet.cell(row=newMaxRow, column = 15)
        # weekNumBefore = datetime.date(tempCell.value).strftime("%V")
        # print(weekNumBefore)
        # print(weekNumToday)
        # #Total num of flights
        # newMaxRow = self.getMaxRow('M') + 1
        # tempCell = self.worksheet.cell(row=newMaxRow, column = 13)
        #  
        # if(weekNumToday != weekNumToday):
        #     average = tempCell.value
        # else:
        #=======================================================================
            
        
           
    def isAlreadyUpdated(self,dateToday):
        maxRow = self.getMaxRow('C')
        lastDateUpdate = self.worksheet.cell(row = maxRow, column = 3).internal_value.strftime('%m/%d/%Y')
       
        if (dateToday == lastDateUpdate):
            return True
        return False
              
    def saveFile(self):
        self.workbook.save(self.filePath)
        
        