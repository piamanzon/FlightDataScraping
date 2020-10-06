'''
Created on Oct 5, 2020

@author: pia.manzon
'''
from openpyxl import workbook #pip install openpyxl
from openpyxl import load_workbook
from openpyxl.chart import (
    LineChart,
    Reference,
)


class Excel():
    def __init__(self,excelPath):
      #   excelPath = r"\\petrochina-ca.com\data\HOME\pia.manzon\YYC_FlightData.xlsx"
        self.filePath = excelPath
        self.workbook = load_workbook(excelPath)
        self.sheet = self.workbook.active
    
    def getMaxRow(self,columnName):
        maxRow = max((c.row for c in self.sheet[columnName] if c.value is not None))
        return maxRow
    
    def inputNewData (self,maxRow,flightDataList,city):
#===============================================================================
        newMaxRow = maxRow
        columnStart = 0
        
        if (city == "Vancouver"):
            columnStart = 3
           
        else:
            columnStart = 7
        
#Update Flight Info
#===============================================================================
        for index in range(len(flightDataList)):
            self.sheet.cell(row=newMaxRow+1, column = columnStart).value = flightDataList[index]['Date']
            self.sheet.cell(row=newMaxRow+1, column = (columnStart+1)).value = flightDataList[index]['Flight']
            self.sheet.cell(row=newMaxRow+1, column = (columnStart+2)).value = flightDataList[index]['From']
            newMaxRow+=1   
#Update Daily Summary
#===============================================================================  

    def updateDailySummaryVan(self,flightVancouverList,flightTorontoList):
        dateToday = flightVancouverList[0]['Date']
        columnStart = 12 
        newMaxRow = self.getMaxRow('l') + 1
        print(newMaxRow)
        self.sheet.cell(row=newMaxRow, column = columnStart).value = dateToday
        self.sheet.cell(row=newMaxRow, column = (columnStart+1)).value = len(flightVancouverList)
        self.sheet.cell(row=newMaxRow, column = (columnStart+2)).value = len(flightTorontoList)
        
    def createGraph(self):
        c1 = LineChart()
        c1.title = "Number of Flight Arrivals at YYC"
        c1.style = 14
        maxRow = self.getMaxRow('L')
        data = Reference(self.sheet, min_col=13, min_row=2, max_col=14, max_row=maxRow)
        c1.add_data(data, titles_from_data=True)
        
        # Style the lines
        s1 = c1.series[0]
        s1.smooth = True # Make the line smooth
        s1.marker.graphicalProperties.solidFill = "4472C4" # Marker filling
        s1.marker.graphicalProperties.line.solidFill = "4472C4" # Marker outline
        
        
        s2 = c1.series[1]
        s2.smooth = True # Make the line smooth
        s2.marker.graphicalProperties.solidFill = "ED7D31" # Marker filling
        s2.marker.graphicalProperties.line.solidFill = "ED7D31" # Marker outline
        
        self.sheet.add_chart(c1, "P2")
                  
    def saveFile(self):
        self.workbook.save(self.filePath)
        
         